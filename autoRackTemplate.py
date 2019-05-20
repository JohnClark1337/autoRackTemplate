from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Color, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from subprocess import check_output
import re
import os
home = os.path.expanduser('~')
doc = ''
wb = Workbook()
sheet = wb.active

#formatting
uFont = Font(color="FFFFFF")
boldFont = Font(bold=True)
align = Alignment(horizontal='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
titleFill = PatternFill(start_color='d9edf3', end_color='d9edf3', fill_type='solid')
uNumFill = PatternFill(start_color='c4bd96', end_color='c4bd96', fill_type='solid')
eCells = PatternFill(start_color="e0e1e2", end_color="e0e1e2", fill_type="solid")

yesList = ["YES", "Y"]
noList = ["NO", "N"]


#Create the cells in excel
def rName(name, r, c):
        sheet.cell(row=r, column=c).value = name.upper()
        sheet.cell(row=r, column=c).font = boldFont
        sheet.cell(row=r, column=c).alignment = align
        sheet.cell(row=r, column=c).fill = titleFill
        sheet.cell(row=r, column=c).border = thin_border


def uCreate(total, c, e=False):
        t = total
        for i in range(2, t+2):
                sheet.cell(row=i, column=c).alignment = align
                sheet.cell(row=i, column=c).border = thin_border
                if e == False:
                        sheet.cell(row=i, column=c).font = uFont
                        sheet.cell(row=i, column=c).fill = uNumFill
                        sheet.cell(row=i, column=c).value = t
                else:
                        sheet.cell(row=i, column=c).fill = eCells
                t-=1
              

#Initial questions for filename
answer =''
fname = ''
print("Welcome to rack template creator.")
while answer != 'Q':
        totalU = 0
        rlist = list()
        l = True
        while l == True:
                fname = str(input("Spreadsheet name: "))
                if fname != '':
                        doc = home + '/Documents/{}.xlsx'.format(fname)
                        if os.path.isfile(doc):
                                t = True
                                while t == True:
                                        ovrw = input("File Already exists!\nOverwrite?(Y/N): ")
                                        if ovrw.upper() in noList:
                                                t = False
                                        elif ovrw.upper() in yesList:
                                                try:
                                                        os.remove(doc)
                                                        t = False
                                                        l = False
                                                except:
                                                        print("Cannot overwrite file")
                                                        t = False
                        else:
                                l = False
        rnum = str(input("What racks are being added?(ex 'A1' or 'A1, B3' or 'A1-A7')"))
        #parsing text if necessary
        if ',' in rnum:
                text = rnum.split(',')
        else:
                text = {rnum}
        for i in text:
                #looking for letter then number(s) together in text
                if '-' not in i:
                        z = re.search(r"[a-zA-Z][0-9]+", i)
                        if z != None:
                                rlist.append(z.string)
                #looking for two sets of letter/number combinations separated by hypen in text
                else:
                        x = re.search(r"[a-zA-Z][0-9]+[-][a-zA-Z]*[0-9]+", i)
                        if x != None:
                                letterList = list()
                                endpoints = list()
                                counter = 0
                                beginning = 0
                                end = 0
                                findRange = x.string.split('-')
                                for r in findRange:
                                        letterList.append(str(''.join(list(filter(str.isalpha, r)))))
                                if letterList[1] is not None and letterList[1] != "":
                                        if letterList[0].upper() == letterList[1].upper():
                                                for r in findRange:
                                                        endpoints.append(int(''.join(list(filter(str.isdigit, r)))))
                                elif letterList[1] is None or letterList[1] == "":
                                        for r in findRange:
                                                endpoints.append(int(''.join(list(filter(str.isdigit, r)))))
                                if endpoints != None:
                                        temp=0
                                        if endpoints[1] > endpoints[0]:
                                                end = endpoints[1]
                                                beginning = endpoints[0]
                                        elif endpoints[0] > endpoints[1]:
                                                beginning = endpoints[1]
                                                end = endpoints[0]
                                        counter = end - beginning
                                        
                                if counter == 0 and (letterList is not None or endpoints is not None):
                                        rlist.append(letterList[0] + endpoints[0])
                                elif counter > 0:
                                        for x in range(beginning, end + 1):
                                                rlist.append(letterList[0] + "{}".format(x))

        #Find out number of U's, create entries in workbook, and save workbook to location
        cloc = 2
        rloc = 1
        if len(rlist) > 0:
                totalU = int(input("\nNumber of U's per rack: "))
                for entry in rlist: 
                        rName(entry, rloc, cloc)
                        sheet.column_dimensions[get_column_letter(cloc)].width = 25
                        uCreate(totalU, cloc - 1)
                        uCreate(totalU, cloc + 1)
                        uCreate(totalU, cloc, True)
                        sheet.column_dimensions[get_column_letter(cloc-1)].width = 5
                        sheet.column_dimensions[get_column_letter(cloc+1)].width = 5
                        cloc += 4
                try:
                        wb.save(doc)
                        check_output(doc, shell=True)
                except Exception as e:
                        print("\nUnable to write file. " + str(e))
        else:
                print("\nProblem with input names.\nNothing Written\n")


        #Ask if user wants to make another template
        answer = ''
        loop = True
        while loop == True:
                print("\nMake another template?\n")
                answer = str(input("(Y/N):"))
                if answer.upper() in noList:
                        print("Have a nice day")
                        loop = False
                        answer = 'Q'
                elif answer.upper() in yesList:
                        print("Making another template.\n\n")
                        loop = False
                else:
                        print("Unknown answer (Y or N)")

        #Resets workbook in preparation of creation a new template
        wb = Workbook()
        sheet = wb.active
